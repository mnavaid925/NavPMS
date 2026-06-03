"""Catalog Management domain services (Module 10).

All state transitions live here, wrapped in ``@transaction.atomic`` with audit
logging via :func:`apps.tenants.services.record_audit`. Mirrors the contracts
service style (perms + numbering + lifecycle + analytics) and adds:

  * a self-contained item / price-change approval workflow,
  * effective-price resolution over volume / contract / date-windowed tiers,
  * real cXML/OCI punch-out orchestration (connectors live in ``punchout.py``), and
  * a CSV/XLSX supplier-upload parser that stages draft items for approval.
"""
import csv
import io
import secrets
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum
from django.urls import reverse
from django.utils import timezone

from apps.core.models import Tenant, set_current_tenant
from apps.portal.models import Notification
from apps.tenants.services import record_audit

from .models import (
    ITEM_ORDERABLE_STATUSES,
    ITEM_SUBMITTABLE_STATUSES,
    UOM_CHOICES,
    CatalogCategory,
    CatalogItem,
    CatalogItemStatusEvent,
    CatalogPriceChangeRequest,
    CatalogPriceTier,
    PunchoutSession,
    SupplierCatalogUpload,
)
from .punchout import (  # noqa: F401  (validate_punchout_url re-exported for models.clean)
    _http_post,
    get_connector,
    validate_punchout_url,
)

# Roles allowed to create/configure/manage the catalog (mirrors contracts).
MANAGE_ROLES = ('tenant_admin', 'procurement_manager', 'buyer')
# Viewing (analytics / read-only) additionally allows approvers.
VIEW_ROLES = MANAGE_ROLES + ('approver',)

PUNCHOUT_SESSION_TTL = timedelta(hours=1)
_UOM_KEYS = {k for k, _ in UOM_CHOICES}


# ---------------------------------------------------------------------------
# Permissions
# ---------------------------------------------------------------------------
def _has_role(user, roles):
    """True if the user holds any of ``roles`` (string slugs)."""
    if not user or not user.is_authenticated:
        return False
    if getattr(user, 'is_superuser', False):
        return True
    if getattr(user, 'is_tenant_admin', False):
        return True
    role = getattr(user, 'role', None)
    if isinstance(role, str):
        role_slug = role
    else:
        role_slug = getattr(role, 'slug', None) or getattr(role, 'name', None)
    return role_slug in roles


def can_manage_catalog(user):
    """May create/configure/approve catalog items and punch-out."""
    return _has_role(user, MANAGE_ROLES)


def can_view_catalog(user):
    """May view the catalog / analytics (managers + approvers)."""
    return _has_role(user, VIEW_ROLES)


def catalog_item_visible_to(user, item):
    """True if ``user`` may view ``item``.

    Internal managers/approvers may view any item in their tenant; a vendor portal
    user may view only items where they are the supplier.
    """
    if not user or not user.is_authenticated:
        return False
    if getattr(user, 'is_vendor_user', False):
        return getattr(user, 'vendor_id', None) == item.vendor_id
    return can_view_catalog(user)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def record_status_event(item, status, user, note='', price_change=None):
    """Append an immutable lifecycle timeline row."""
    return CatalogItemStatusEvent.all_objects.create(
        tenant=item.tenant, item=item, price_change=price_change, status=status,
        note=(note or '')[:255], actor=user,
    )


def _notify(user, item, *, category, priority, title, message):
    """Create a portal Notification linking to a catalog item."""
    if not user:
        return None
    try:
        link = reverse('catalog:item_detail', kwargs={'pk': item.pk})
    except Exception:
        link = ''
    return Notification.all_objects.create(
        tenant=item.tenant, user=user, category=category, priority=priority,
        title=title[:160], message=message, link_url=link,
    )


def _notify_managers(item, *, category, priority, title, message, exclude=None):
    """Notify every manage-role user in the tenant (the approval queue)."""
    from apps.accounts.models import User
    users = User.objects.filter(tenant=item.tenant, is_active=True).filter(
        Q(is_tenant_admin=True) | Q(role__in=MANAGE_ROLES))
    for u in users:
        if exclude and u.pk == exclude.pk:
            continue
        _notify(u, item, category=category, priority=priority,
                title=title, message=message)


# ---------------------------------------------------------------------------
# Numbering
# ---------------------------------------------------------------------------
def next_item_number(tenant):
    """Generate the next gap-free ``CAT-<SLUG>-NNNNN`` for ``tenant``."""
    slug = (tenant.slug or str(tenant.pk))[:6].upper()
    prefix = f'CAT-{slug}-'
    last = (
        CatalogItem.all_objects
        .filter(tenant=tenant, item_number__startswith=prefix)
        .order_by('-item_number')
        .first()
    )
    seq = 1
    if last:
        try:
            seq = int(last.item_number.rsplit('-', 1)[1]) + 1
        except (IndexError, ValueError):
            seq = 1
    number = f'{prefix}{seq:05d}'
    while CatalogItem.all_objects.filter(
            tenant=tenant, item_number=number).exists():
        seq += 1
        number = f'{prefix}{seq:05d}'
    return number


def next_price_change_number(item):
    """Return the next ``<item_number>-PC0N`` price-change label."""
    n = CatalogPriceChangeRequest.all_objects.filter(item=item).count() + 1
    return f'{item.item_number}-PC{n:02d}'


# ---------------------------------------------------------------------------
# Item creation & approval workflow
# ---------------------------------------------------------------------------
def create_item(*, tenant, user, **fields):
    """Create a draft CatalogItem with an auto-assigned, collision-safe number."""
    last_exc = None
    for _attempt in range(5):
        try:
            with transaction.atomic():
                Tenant.objects.select_for_update().get(pk=tenant.pk)
                item = CatalogItem.all_objects.create(
                    tenant=tenant,
                    item_number=next_item_number(tenant),
                    status='draft',
                    created_by=user,
                    **fields,
                )
                record_status_event(item, 'draft', user, 'Item created')
                record_audit(
                    tenant, user, 'catalog.item.created',
                    target_type='CatalogItem', target_id=str(item.id),
                    message=f'{item.item_number}: {item.name}',
                )
            return item
        except IntegrityError as exc:
            last_exc = exc
    raise last_exc


def validate_item_for_submission(item):
    """Raise ``ValidationError`` unless the item is ready to submit for approval."""
    errors = []
    if not (item.name or '').strip():
        errors.append('Give the item a name before submitting.')
    if item.base_price is None or item.base_price < 0:
        errors.append('Set a non-negative base price.')
    if item.source == 'supplier' and not item.vendor_id:
        errors.append('Choose a supplier for a supplier-sourced item.')
    if errors:
        raise ValidationError(errors)
    return []


def submit_item_for_approval(item, user):
    """Transition draft/rejected → pending_approval."""
    with transaction.atomic():
        if item.status not in ITEM_SUBMITTABLE_STATUSES:
            raise ValidationError('Only draft or rejected items can be submitted.')
        validate_item_for_submission(item)
        item.status = 'pending_approval'
        item.submitted_at = timezone.now()
        item.rejection_reason = ''
        item.save(update_fields=[
            'status', 'submitted_at', 'rejection_reason', 'updated_at'])
        record_status_event(item, 'pending_approval', user, 'Submitted for approval')
        record_audit(
            item.tenant, user, 'catalog.item.submitted',
            target_type='CatalogItem', target_id=str(item.id),
            message=f'{item.item_number} submitted for approval',
        )
        _notify_managers(
            item, category='approval', priority='normal',
            title=f'Catalog item awaiting approval: {item.item_number}',
            message=f'“{item.name}” was submitted for review.', exclude=user,
        )
    return item


def approve_item(item, user, note=''):
    """Approve a pending item, making it orderable."""
    with transaction.atomic():
        if item.status != 'pending_approval':
            raise ValidationError('Only items awaiting approval can be approved.')
        item.status = 'approved'
        item.approved_at = timezone.now()
        item.approved_by = user
        item.save(update_fields=['status', 'approved_at', 'approved_by', 'updated_at'])
        record_status_event(item, 'approved', user, note or 'Approved')
        record_audit(
            item.tenant, user, 'catalog.item.approved',
            target_type='CatalogItem', target_id=str(item.id),
            message=f'{item.item_number} approved',
        )
        if item.created_by_id and item.created_by_id != getattr(user, 'pk', None):
            _notify(item.created_by, item, category='info', priority='normal',
                    title=f'Catalog item approved: {item.item_number}',
                    message=f'“{item.name}” is now in the catalog.')
    return item


def reject_item(item, user, reason):
    """Reject a pending item; it returns to an editable state for revision."""
    with transaction.atomic():
        if item.status != 'pending_approval':
            raise ValidationError('Only items awaiting approval can be rejected.')
        item.status = 'rejected'
        item.rejected_at = timezone.now()
        item.rejection_reason = (reason or '').strip()[:255]
        item.save(update_fields=[
            'status', 'rejected_at', 'rejection_reason', 'updated_at'])
        record_status_event(item, 'rejected', user, item.rejection_reason)
        record_audit(
            item.tenant, user, 'catalog.item.rejected', level='warning',
            target_type='CatalogItem', target_id=str(item.id),
            message=f'{item.item_number} rejected: {item.rejection_reason}'[:255],
        )
        if item.created_by_id and item.created_by_id != getattr(user, 'pk', None):
            _notify(item.created_by, item, category='approval', priority='high',
                    title=f'Catalog item rejected: {item.item_number}',
                    message=f'“{item.name}” needs changes: {item.rejection_reason}')
    return item


def retire_item(item, user, reason=''):
    """Retire an approved item (no longer orderable, kept for history)."""
    with transaction.atomic():
        if item.status != 'approved':
            raise ValidationError('Only approved items can be retired.')
        item.status = 'retired'
        item.retired_at = timezone.now()
        item.is_active = False
        item.save(update_fields=['status', 'retired_at', 'is_active', 'updated_at'])
        record_status_event(item, 'retired', user, (reason or 'Retired')[:255])
        record_audit(
            item.tenant, user, 'catalog.item.retired', level='warning',
            target_type='CatalogItem', target_id=str(item.id),
            message=f'{item.item_number} retired',
        )
    return item


def archive_item(item, user):
    """Archive a non-open item (drop it out of the working set)."""
    with transaction.atomic():
        if item.status in ('pending_approval',):
            raise ValidationError('Resolve the pending approval before archiving.')
        item.status = 'archived'
        item.is_active = False
        item.save(update_fields=['status', 'is_active', 'updated_at'])
        record_status_event(item, 'archived', user, 'Archived')
        record_audit(
            item.tenant, user, 'catalog.item.archived', level='warning',
            target_type='CatalogItem', target_id=str(item.id),
            message=f'{item.item_number} archived',
        )
    return item


# ---------------------------------------------------------------------------
# Pricing & tier resolution
# ---------------------------------------------------------------------------
def current_tiers(item, on_date=None):
    """Active price tiers for ``item`` whose effective window contains ``on_date``."""
    on_date = on_date or timezone.localdate()
    qs = CatalogPriceTier.all_objects.filter(item=item, is_active=True)
    return [
        t for t in qs
        if (t.effective_from is None or t.effective_from <= on_date)
        and (t.effective_to is None or on_date <= t.effective_to)
    ]


def resolve_price(item, *, qty=None, on_date=None, contract=None):
    """Resolve the best (lowest) current unit price for ``item`` at ``qty``.

    Considers active, in-window tiers whose ``min_quantity`` is satisfied; a tier
    tied to ``contract`` is preferred when a contract is supplied. Falls back to
    ``item.base_price`` when no tier applies.
    """
    if qty is None:
        qty = item.min_order_qty or Decimal('1')
    try:
        qty = Decimal(str(qty))
    except (InvalidOperation, ValueError):
        qty = Decimal('1')

    candidates = [t for t in current_tiers(item, on_date) if t.min_quantity <= qty]
    if contract is not None:
        contract_tiers = [t for t in candidates if t.contract_id == contract.pk]
        if contract_tiers:
            candidates = contract_tiers
    if not candidates:
        return item.base_price or Decimal('0.0000')
    return min(t.unit_price for t in candidates)


def _serialize_tiers(item):
    """Snapshot an item's current tiers (for the price-change audit trail)."""
    out = []
    for t in CatalogPriceTier.all_objects.filter(item=item):
        out.append({
            'min_quantity': str(t.min_quantity),
            'unit_price': str(t.unit_price),
            'tier_type': t.tier_type,
            'effective_from': t.effective_from.isoformat() if t.effective_from else None,
            'effective_to': t.effective_to.isoformat() if t.effective_to else None,
        })
    return out


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return None


def submit_price_change(pc, user):
    """Transition a price-change request draft → pending_approval."""
    with transaction.atomic():
        if pc.status != 'draft':
            raise ValidationError('Only draft price-change requests can be submitted.')
        pc.status = 'pending_approval'
        pc.submitted_at = timezone.now()
        pc.save(update_fields=['status', 'submitted_at', 'updated_at'])
        record_status_event(pc.item, 'pending_approval', user,
                            f'Price change {pc.request_number} submitted', price_change=pc)
        record_audit(
            pc.tenant, user, 'catalog.price_change.submitted',
            target_type='CatalogPriceChangeRequest', target_id=str(pc.id),
            message=f'{pc.request_number} submitted',
        )
        _notify_managers(
            pc.item, category='approval', priority='normal',
            title=f'Price change awaiting approval: {pc.item.item_number}',
            message=f'Review price change {pc.request_number}.', exclude=user)
    return pc


def apply_price_change(pc, user):
    """Approve a price-change request, applying the new base price and/or tiers."""
    with transaction.atomic():
        item = CatalogItem.all_objects.select_for_update().get(pk=pc.item_id)
        if not pc.is_editable and pc.status != 'pending_approval':
            raise ValidationError('This price-change request has already been resolved.')
        if item.status != 'approved':
            raise ValidationError('Only an approved item can have a price change applied.')

        pc.prev_base_price = item.base_price
        pc.prev_tiers = _serialize_tiers(item)

        if pc.change_type in ('base', 'both') and pc.new_base_price is not None:
            item.base_price = pc.new_base_price
            item.save(update_fields=['base_price', 'updated_at'])

        if pc.change_type in ('tiers', 'both'):
            CatalogPriceTier.all_objects.filter(item=item).delete()
            for row in (pc.proposed_tiers or []):
                CatalogPriceTier.all_objects.create(
                    tenant=item.tenant, item=item,
                    tier_type=row.get('tier_type', 'volume'),
                    min_quantity=Decimal(str(row.get('min_quantity', '1'))),
                    unit_price=Decimal(str(row.get('unit_price', '0'))),
                    effective_from=_parse_date(row.get('effective_from')),
                    effective_to=_parse_date(row.get('effective_to')),
                )

        pc.status = 'approved'
        pc.decided_at = timezone.now()
        pc.decided_by = user
        pc.save(update_fields=[
            'status', 'decided_at', 'decided_by', 'prev_base_price', 'prev_tiers',
            'updated_at'])
        record_status_event(item, 'approved', user,
                            f'Price change {pc.request_number} applied', price_change=pc)
        record_audit(
            pc.tenant, user, 'catalog.price_change.applied',
            target_type='CatalogPriceChangeRequest', target_id=str(pc.id),
            message=f'{pc.request_number} applied to {item.item_number}',
            payload={
                'item_id': item.id,
                'new_base_price': str(pc.new_base_price) if pc.new_base_price is not None else None,
                'tier_count': len(pc.proposed_tiers or []),
            },
        )
    return pc


def reject_price_change(pc, user, reason=''):
    """Reject a pending price-change request."""
    with transaction.atomic():
        if pc.status not in ('draft', 'pending_approval'):
            raise ValidationError('This price-change request has already been resolved.')
        pc.status = 'rejected'
        pc.decided_at = timezone.now()
        pc.decided_by = user
        pc.decision_note = (reason or '').strip()[:255]
        pc.save(update_fields=[
            'status', 'decided_at', 'decided_by', 'decision_note', 'updated_at'])
        record_audit(
            pc.tenant, user, 'catalog.price_change.rejected', level='warning',
            target_type='CatalogPriceChangeRequest', target_id=str(pc.id),
            message=f'{pc.request_number} rejected',
        )
    return pc


def cancel_price_change(pc, user):
    """Cancel a draft/pending price-change request without touching the item."""
    with transaction.atomic():
        if pc.status not in ('draft', 'pending_approval'):
            raise ValidationError('This price-change request has already been resolved.')
        pc.status = 'cancelled'
        pc.save(update_fields=['status', 'updated_at'])
        record_audit(
            pc.tenant, user, 'catalog.price_change.cancelled', level='warning',
            target_type='CatalogPriceChangeRequest', target_id=str(pc.id),
            message=f'{pc.request_number} cancelled',
        )
    return pc


# ---------------------------------------------------------------------------
# Punch-out orchestration (real cXML / OCI)
# ---------------------------------------------------------------------------
def start_punchout(config, user, *, build_return_url, requisition=None):
    """Open a punch-out session and (for cXML) perform the server-side setup POST.

    ``build_return_url`` is a ``callable(return_token) -> absolute_url`` (the view
    supplies ``request.build_absolute_uri`` over the inbound endpoint) — the token
    is minted here so the return URL can embed it. For browser-mediated protocols
    (OCI/loopback) the bridge view renders the auto-POST form; for cXML the
    StartPage is fetched here and SSRF-revalidated.
    """
    with transaction.atomic():
        session = PunchoutSession.all_objects.create(
            tenant=config.tenant, config=config, vendor=config.vendor,
            buyer_cookie=secrets.token_urlsafe(24),
            return_token=secrets.token_urlsafe(32),
            started_by=user, requisition=requisition,
            expires_at=timezone.now() + PUNCHOUT_SESSION_TTL,
        )
        return_url = build_return_url(session.return_token)
        connector = get_connector(config)
        descriptor = connector.build_setup(
            config=config, session=session, return_url=return_url)

        if descriptor.server_post:
            validate_punchout_url(descriptor.url)  # SSRF guard before the call
            try:
                resp = _http_post(descriptor.url, descriptor.body, descriptor.headers)
            except Exception as exc:  # network / transport failure
                session.status = 'failed'
                session.error_message = f'Setup request failed: {exc}'[:255]
                session.save(update_fields=['status', 'error_message', 'updated_at'])
                raise ValidationError(f'Punch-out setup failed: {exc}')
            start = connector.parse_setup_response(getattr(resp, 'text', ''))
            if not start:
                session.status = 'failed'
                session.error_message = 'Supplier did not return a start page.'
                session.save(update_fields=['status', 'error_message', 'updated_at'])
                raise ValidationError('Supplier did not return a start page.')
            validate_punchout_url(start)  # SSRF guard on the supplier-returned URL
            session.start_page_url = start
            session.status = 'redirected'
            session.redirected_at = timezone.now()
            session.save(update_fields=[
                'start_page_url', 'status', 'redirected_at', 'updated_at'])
        else:
            session.start_page_url = descriptor.url
            session.save(update_fields=['start_page_url', 'updated_at'])

        record_audit(
            config.tenant, user, 'catalog.punchout.started',
            target_type='PunchoutSession', target_id=str(session.id),
            message=f'Punch-out to {config.vendor.legal_name} ({config.protocol})',
        )
    return session


def receive_punchout_order(request, session):
    """Validate and parse an inbound PunchOutOrderMessage / OCI return into a cart.

    WARNING: this endpoint is necessarily CSRF-exempt (a cross-site supplier POST),
    so it is authenticated by the unguessable ``return_token`` (resolved by the
    view) AND the connector's inbound credential check. Never log the secret.
    """
    connector = get_connector(session.config)
    if session.is_expired:
        session.status = 'expired'
        session.save(update_fields=['status', 'updated_at'])
        raise ValidationError('This punch-out session has expired.')
    if session.status == 'returned':
        raise ValidationError('This punch-out cart has already been received.')

    if not connector.authenticate_inbound(
            request=request, config=session.config, session=session):
        record_audit(
            session.tenant, session.started_by, 'catalog.punchout.auth_failed',
            level='warning', target_type='PunchoutSession', target_id=str(session.id),
            message=f'Inbound punch-out authentication failed for session {session.id}',
        )
        raise ValidationError('Punch-out authentication failed.')

    cart = connector.parse_cart(
        request=request, config=session.config, session=session)
    if not cart.ok:
        session.status = 'failed'
        session.error_message = (cart.message or 'Could not parse the cart.')[:255]
        session.save(update_fields=['status', 'error_message', 'updated_at'])
        raise ValidationError(session.error_message)

    session.cart_data = [
        {
            'name': line['name'], 'sku': line.get('sku', ''),
            'quantity': str(line['quantity']), 'unit_price': str(line['unit_price']),
            'currency': line.get('currency', 'USD'), 'uom': line.get('uom', 'each'),
        }
        for line in cart.lines
    ]
    session.status = 'returned'
    session.returned_at = timezone.now()
    session.save(update_fields=['cart_data', 'status', 'returned_at', 'updated_at'])
    record_audit(
        session.tenant, session.started_by, 'catalog.punchout.returned',
        target_type='PunchoutSession', target_id=str(session.id),
        message=f'Punch-out cart received: {len(session.cart_data)} line(s)',
    )
    return session


def cart_to_requisition_lines(session, requisition, user):
    """Convert a returned punch-out cart into requisition lines."""
    from apps.requisitions.models import RequisitionLine
    created = 0
    with transaction.atomic():
        for line in (session.cart_data or []):
            RequisitionLine.all_objects.create(
                tenant=session.tenant, requisition=requisition,
                description=line['name'][:255],
                quantity=Decimal(str(line.get('quantity', '1'))),
                unit=line.get('uom', 'each')[:30],
                unit_price=Decimal(str(line.get('unit_price', '0'))),
            )
            created += 1
        if hasattr(requisition, 'recalc_total'):
            requisition.recalc_total()
        record_audit(
            session.tenant, user, 'catalog.punchout.to_requisition',
            target_type='Requisition', target_id=str(requisition.id),
            message=f'{created} punch-out line(s) added to {requisition.number}',
        )
    return created


def cart_to_staged_items(session, user):
    """Stage a returned punch-out cart as draft supplier catalog items."""
    created = []
    for line in (session.cart_data or []):
        item = create_item(
            tenant=session.tenant, user=user, name=line['name'][:200],
            source='supplier', vendor=session.vendor, sku=line.get('sku', '')[:60],
            uom=(line.get('uom') if line.get('uom') in _UOM_KEYS else 'each'),
            currency=line.get('currency', 'USD')[:3],
            base_price=Decimal(str(line.get('unit_price', '0'))),
            source_session=session,
        )
        created.append(item)
    return created


# ---------------------------------------------------------------------------
# Supplier catalog upload — parse & ingest
# ---------------------------------------------------------------------------
_REQUIRED_COLUMNS = ('name', 'base_price')


def parse_csv(file):
    """Yield header-mapped dict rows from a CSV upload."""
    file.seek(0)
    text = file.read()
    if isinstance(text, bytes):
        text = text.decode('utf-8-sig', 'replace')
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        yield {(k or '').strip().lower(): (v or '').strip() for k, v in row.items()}


def parse_xlsx(file):
    """Yield header-mapped dict rows from an XLSX upload."""
    from openpyxl import load_workbook
    file.seek(0)
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    keys = [str(h).strip().lower() if h is not None else '' for h in header]
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        record = {}
        for i, key in enumerate(keys):
            if not key:
                continue
            val = values[i] if i < len(values) else None
            record[key] = '' if val is None else str(val).strip()
        yield record


def validate_catalog_row(row, tenant, vendor, default_category=None):
    """Validate a parsed upload row. Returns ``(cleaned | None, errors)``."""
    errors = []
    name = (row.get('name') or '').strip()
    if not name:
        errors.append({'field': 'name', 'message': 'Name is required.'})

    raw_price = (row.get('base_price') or row.get('price') or '').strip()
    try:
        price = Decimal(raw_price) if raw_price else Decimal('0')
        if price < 0:
            errors.append({'field': 'base_price', 'message': 'Price must be ≥ 0.'})
    except (InvalidOperation, ValueError):
        price = Decimal('0')
        errors.append({'field': 'base_price', 'message': f'Invalid price “{raw_price}”.'})

    uom = (row.get('uom') or 'each').strip().lower()
    if uom not in _UOM_KEYS:
        uom = 'each'

    category = default_category
    code = (row.get('category_code') or row.get('category') or '').strip()
    if code:
        match = CatalogCategory.all_objects.filter(tenant=tenant, code__iexact=code).first()
        if match:
            category = match

    try:
        moq = Decimal(row.get('min_order_qty') or '1')
    except (InvalidOperation, ValueError):
        moq = Decimal('1')
    try:
        lead = int(float(row.get('lead_time_days') or '0'))
    except (TypeError, ValueError):
        lead = 0

    if errors:
        return None, errors
    return {
        'name': name[:200],
        'description': (row.get('description') or '')[:2000],
        'sku': (row.get('sku') or '')[:60],
        'manufacturer_part_number': (row.get('manufacturer_part_number') or row.get('mpn') or '')[:80],
        'keywords': (row.get('keywords') or '')[:255],
        'base_price': price,
        'uom': uom,
        'min_order_qty': moq,
        'lead_time_days': max(lead, 0),
        'category': category,
    }, []


def process_catalog_upload(upload, user):
    """Parse a supplier upload and stage valid rows as draft catalog items."""
    with transaction.atomic():
        upload.status = 'processing'
        upload.save(update_fields=['status', 'updated_at'])

        name = (upload.original_filename or upload.file.name or '').lower()
        parser = parse_xlsx if name.endswith('.xlsx') else parse_csv

        imported = 0
        errors = []
        row_num = 0
        try:
            for row_num, row in enumerate(parser(upload.file), start=1):
                cleaned, row_errors = validate_catalog_row(
                    row, upload.tenant, upload.vendor, upload.category)
                if row_errors:
                    for e in row_errors:
                        errors.append({'row': row_num, **e})
                    continue
                create_item(
                    tenant=upload.tenant, user=user, source='supplier',
                    vendor=upload.vendor, source_upload=upload, **cleaned,
                )
                imported += 1
        except Exception as exc:  # malformed file
            upload.status = 'failed'
            upload.error_log = errors + [{'row': row_num, 'field': 'file',
                                          'message': f'Could not read file: {exc}'}]
            upload.row_count = row_num
            upload.imported_count = imported
            upload.error_count = len(upload.error_log)
            upload.processed_at = timezone.now()
            upload.save(update_fields=[
                'status', 'error_log', 'row_count', 'imported_count',
                'error_count', 'processed_at', 'updated_at'])
            record_audit(
                upload.tenant, user, 'catalog.upload.failed', level='warning',
                target_type='SupplierCatalogUpload', target_id=str(upload.id),
                message=f'Catalog upload failed: {exc}'[:255])
            return upload

        upload.row_count = row_num
        upload.imported_count = imported
        upload.error_count = len(errors)
        upload.error_log = errors
        upload.processed_at = timezone.now()
        if imported and errors:
            upload.status = 'partially_imported'
        elif imported:
            upload.status = 'imported'
        else:
            upload.status = 'failed'
        upload.save(update_fields=[
            'row_count', 'imported_count', 'error_count', 'error_log',
            'processed_at', 'status', 'updated_at'])
        record_audit(
            upload.tenant, user, 'catalog.upload.processed',
            target_type='SupplierCatalogUpload', target_id=str(upload.id),
            message=f'Upload staged {imported} item(s), {len(errors)} error(s)',
        )
    return upload


# ---------------------------------------------------------------------------
# Analytics
# ---------------------------------------------------------------------------
def tenant_catalog_metrics(tenant):
    """Aggregate catalog KPIs for the analytics dashboard."""
    qs = CatalogItem.objects.filter(tenant=tenant)
    by_status = dict(qs.values_list('status').annotate(n=Count('id')))
    by_source = dict(qs.values_list('source').annotate(n=Count('id')))

    sessions = PunchoutSession.objects.filter(tenant=tenant)
    uploads = SupplierCatalogUpload.objects.filter(tenant=tenant)

    top_categories = list(
        qs.filter(status='approved', category__isnull=False)
        .values('category__name')
        .annotate(n=Count('id')).order_by('-n')[:5]
    )
    top_vendors = list(
        qs.filter(status='approved', vendor__isnull=False)
        .values('vendor__legal_name')
        .annotate(n=Count('id')).order_by('-n')[:5]
    )

    return {
        'total_items': qs.count(),
        'by_status': by_status,
        'draft': by_status.get('draft', 0),
        'pending_approval': by_status.get('pending_approval', 0),
        'approved': by_status.get('approved', 0),
        'rejected': by_status.get('rejected', 0),
        'retired': by_status.get('retired', 0),
        'archived': by_status.get('archived', 0),
        'internal': by_source.get('internal', 0),
        'supplier': by_source.get('supplier', 0),
        'active_tiers': CatalogPriceTier.objects.filter(
            item__tenant=tenant, is_active=True).count(),
        'pending_price_changes': CatalogPriceChangeRequest.objects.filter(
            tenant=tenant, status='pending_approval').count(),
        'punchout_sessions': sessions.count(),
        'punchout_returned': sessions.filter(status='returned').count(),
        'open_uploads': uploads.filter(status__in=('pending', 'processing')).count(),
        'top_categories': top_categories,
        'top_vendors': top_vendors,
    }


def catalog_item_analytics(item):
    """Per-item analytics: pricing, tiers, approval and change history."""
    tiers = CatalogPriceTier.all_objects.filter(item=item)
    return {
        'tier_count': tiers.count(),
        'current_tier_count': len(current_tiers(item)),
        'effective_price': resolve_price(item, qty=item.min_order_qty),
        'base_price': item.base_price,
        'price_change_count': CatalogPriceChangeRequest.all_objects.filter(item=item).count(),
        'status_event_count': CatalogItemStatusEvent.all_objects.filter(item=item).count(),
        'is_orderable': item.status in ITEM_ORDERABLE_STATUSES,
    }
